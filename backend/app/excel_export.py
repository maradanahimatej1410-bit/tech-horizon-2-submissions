import io
import json
import csv
from datetime import datetime
# pyrefly: ignore [missing-import]
from openpyxl import Workbook
# pyrefly: ignore [missing-import]
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# pyrefly: ignore [missing-import]
from openpyxl.utils import get_column_letter
# pyrefly: ignore [missing-import]
from sqlalchemy.orm import Session
from . import models

def format_excel_sheet(ws):
    # Font styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Fill colors
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")  # IEEE Navy Blue
    
    # Freeze pane (freeze first row)
    ws.freeze_panes = "A2"
    
    # Enable auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    
    # Apply styling to headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

    # Format data cells
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            # Date formatting and alignment
            if isinstance(cell.value, datetime):
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                cell.alignment = Alignment(horizontal="center")
            elif str(cell.value).startswith("TH26-"):
                cell.alignment = Alignment(horizontal="center")
            elif cell.value in ["Yes", "No", "✅ Completed", "🟡 Pending", "❌ Not Submitted", "Approved", "Rejected", "Under Review", "Needs Correction"]:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                # Give headers slightly more room
                max_len = max(max_len, len(val) + 4)
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

def generate_registration_excel(db: Session) -> io.BytesIO:
    teams = db.query(models.Team).all()
    wb = Workbook()
    
    # Redesigned tabular sheet with each participant in one separate row
    ws = wb.active
    ws.title = "REGISTRATIONS"
    ws.append([
        "Team ID", "Team Name", "Participant Name", "Gender", "Email", "WhatsApp Number", 
        "College / Company", "Designation", "Graduation Year / Semester", "State", "City", 
        "IEEE Member", "Selected Theme", "Accommodation Required"
    ])
    for team in teams:
        for p in sorted(team.participants, key=lambda x: x.member_index):
            ws.append([
                team.id,
                team.team_name,
                p.name,
                p.gender or "N/A",
                p.email,
                p.whatsapp,
                p.college,
                p.designation,
                p.grad_year_sem,
                p.state,
                p.city,
                "Yes" if p.is_ieee_member else "No",
                team.selected_theme,
                "Yes" if team.accommodation_required else "No"
            ])
    format_excel_sheet(ws)
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def generate_project_idea_excel(db: Session) -> io.BytesIO:
    teams = db.query(models.Team).filter(models.Team.idea_submission_status == "Submitted").all()
    wb = Workbook()
    
    ws = wb.active
    ws.title = "PROJECT IDEAS"
    ws.append([
        "Team ID", "Team Name", "Selected Theme", "Team Lead Name", "Gender", 
        "Team Lead Email", "Team Lead WhatsApp Number", "College / Company", "Graduation Year / Semester", 
        "State", "City", "Submitted Google Drive Link", "Submission Date", "Verification Status", "Organizer Remarks"
    ])
    
    for team in teams:
        lead = next((p for p in team.participants if p.member_index == 1), None)
        ws.append([
            team.id,
            team.team_name,
            team.selected_theme,
            lead.name if lead else "",
            (lead.gender or "N/A") if lead else "",
            lead.email if lead else "",
            lead.whatsapp if lead else "",
            lead.college if lead else "",
            lead.grad_year_sem if lead else "",
            lead.state if lead else "",
            lead.city if lead else "",
            team.project_idea_link or "",
            team.project_idea_date or "",
            team.project_idea_verification_status or "Pending Review",
            team.project_idea_remarks or ""
        ])
        
    format_excel_sheet(ws)
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def generate_registration_csv(db: Session) -> str:
    teams = db.query(models.Team).all()
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        "Team ID", "Team Name", "Theme", "Team Lead", "Lead Email", 
        "Lead WhatsApp", "College", "Accommodation", "Registration Status", "Created At"
    ])
    
    for team in teams:
        lead = next((p for p in team.participants if p.member_index == 1), None)
        accom = "Yes" if team.accommodation_required else "No"
        created_at = team.created_at.strftime("%Y-%m-%d %H:%M:%S") if team.created_at else ""
        
        writer.writerow([
            team.id,
            team.team_name,
            team.selected_theme,
            lead.name if lead else "",
            lead.email if lead else "",
            lead.whatsapp if lead else "",
            lead.college if lead else "",
            accom,
            team.registration_status,
            created_at
        ])
        
    return output.getvalue()

def generate_project_idea_csv(db: Session) -> str:
    teams = db.query(models.Team).filter(models.Team.idea_submission_status == "Submitted").all()
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        "Submission Date", "Submission Time", "Team ID", "Team Name", "Theme", 
        "College", "Project Idea Drive Link", "Submission Status", "Verification Status", "Organizer Remarks"
    ])
    
    for team in teams:
        lead = next((p for p in team.participants if p.member_index == 1), None)
        writer.writerow([
            team.project_idea_date or "",
            team.project_idea_time or "",
            team.id,
            team.team_name,
            team.selected_theme,
            lead.college if lead else "",
            team.project_idea_link or "",
            team.idea_submission_status,
            team.project_idea_verification_status or "Pending Review",
            team.project_idea_remarks or ""
        ])
        
    return output.getvalue()

def generate_json_backup(db: Session) -> str:
    teams = db.query(models.Team).all()
    backup_data = []
    for t in teams:
        t_data = {
            "id": t.id,
            "team_name": t.team_name,
            "selected_theme": t.selected_theme,
            "accommodation_required": t.accommodation_required,
            "referral_source": t.referral_source,
            "additional_comments": t.additional_comments,
            "registration_status": t.registration_status,
            "verification_status": t.verification_status,
            "ticket_status": t.ticket_status,
            "idea_submission_status": t.idea_submission_status,
            "project_idea_link": t.project_idea_link,
            "project_idea_date": t.project_idea_date,
            "project_idea_time": t.project_idea_time,
            "project_idea_verification_status": t.project_idea_verification_status,
            "project_idea_remarks": t.project_idea_remarks,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "participants": [
                {
                    "name": p.name,
                    "email": p.email,
                    "whatsapp": p.whatsapp,
                    "linkedin": p.linkedin,
                    "college": p.college,
                    "designation": p.designation,
                    "grad_year_sem": p.grad_year_sem,
                    "state": p.state,
                    "city": p.city,
                    "is_ieee_member": p.is_ieee_member,
                    "ieee_id_proof_link": p.ieee_id_proof_link,
                    "college_id_proof_link": p.college_id_proof_link,
                    "member_index": p.member_index
                } for p in t.participants
            ],
            "verification": {
                "team_lead_name": t.verification.team_lead_name,
                "company_college": t.verification.company_college,
                "team_name": t.verification.team_name,
                "team_size": t.verification.team_size,
                "ieee_members_count": t.verification.ieee_members_count,
                "non_ieee_members_count": t.verification.non_ieee_members_count,
                "member1_link": t.verification.member1_link,
                "member2_link": t.verification.member2_link,
                "member3_link": t.verification.member3_link,
                "member4_link": t.verification.member4_link,
                "member5_link": t.verification.member5_link,
                "member6_link": t.verification.member6_link,
                "remarks": t.verification.remarks
            } if t.verification else None,
            "notes": [
                {
                    "note_text": n.note_text,
                    "created_at": n.created_at.isoformat() if n.created_at else None
                } for n in t.notes
            ]
        }
        backup_data.append(t_data)
    
    return json.dumps(backup_data, indent=4)
